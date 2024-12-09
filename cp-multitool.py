#!/usr/bin/env python3
# 2024 - DME
# Script to fetch objects, policies from Check Point Management and putting them into a csv / xls file
# https://www.buymeacoffee.com/leinadeuntdomus
#################################################################################################
from cpapi import APIClient, APIClientArgs
import sys
import argparse
import csv
import logging
import datetime
import getpass
try:
    import pandas
except:
    print("Module 'pandas' not found, therefore exports into .xls are not possible.")
    pds = False
else:
    pds = True
    import pandas
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.utils import get_column_letter

#################################################################################################
#
# USAGE: execute script with python according to informations given by help. Possible usages:
# you can set environment variables in lines below, so you don´t have to type them everytime, you execute the script. (Sensitive data should not be saved (like password)
# 
# python3.exe export-objects_subparser.py (--api_server 192.168.0.2 - can be declared by variable "api_server") --api_pwd XuOzvnqDP20xG1LZ48hm5A== -f "IoT Next" show groups
# show groups are the commands here, which then are fired towards API as "show-groups" - see "supported commands" to find out, what is currently implemented
#
#################################################################################################
# set args (will overriden by command)
# static environment variables
# api_server    =   ""
# api_user      =   "" 
# api_pwd       =   ""
# api_context   =   None
# exportselect  =   ""  #"name, uid, groups"
# cmd           =   ""  #"test, show, export"
# choice        =   "hosts"
# cfilter       =   "Check Point"
# 
#################################################################################################


supported_commands = ["test",
                      "show-hosts",
                      "show-groups",
                      "show-networks",
                      "show-services",
                      "show-access-rulebase",
                      "show-packages",
                      "show-unused-objects",
                      "show-updatable-objects-repository-content"]

parser = argparse.ArgumentParser()
parser.add_argument("-H", "--api_server", help="Target Host (CP Management Server)")
parser.add_argument("-U", "--api_user", help="API User")
parser.add_argument("-K", "--key", help="invoke, that password value is an API Key", action="store_true", default=False)
parser.add_argument("-P", "--api_pwd", help="API Users Password, if using API Key, use this without a user")
parser.add_argument("-C", "--api_context", help="If SmartCloud-1 is used, enter context information here (i.e. bhkjnkm-knjhbas-d32424b/web_api) - for On Prem enter \"-C none\"")
parser.add_argument("-p", "--polname", help="when using \"show-access-rulebase\" a policy name must be given, using this argument")
parser.add_argument("-f", "--filter", dest="cfilter" ,help="filter by string (if applicable)")
parser.add_argument("-o", "--outfile", help="filename, to save the output in")
parser.add_argument("-e", "--exportselect", help="choose which fields to export to csv (define as string with quotes: \"name, ip_address, uid, groups\")")
parser.add_argument("-v", "--verbose", help="Run Script with informational (-v) or debugging (-vv) logging output. For troubleshooting purposes.", action="count")

subparser = parser.add_subparsers(required=True, dest="cmd", help="Tell what to do (test || show || export)")
parser_show = subparser.add_parser("show", help="Print given information / objects and their properties (use with caution)")
parser_export = subparser.add_parser("export", help="Save output of given information / objects and their properties to a file")
parser_test = subparser.add_parser("test", help="Basic connectivity test and (if successful) fetch some information from server")
parser_show.add_argument(dest="choice", choices=["hosts","groups","networks","services","policy","unused-objects","packages","uo-content"])
parser_export.add_argument(dest="choice", choices=["hosts","groups","networks","services","policy","unused-objects","uo-content"])

args = parser.parse_args()


#################################################################################################
# ADDING DEBUG MODE                                                                             #
#################################################################################################
if args.verbose == 1:
    logging.basicConfig(level=logging.INFO)
    logging.debug("################## Starting - With informational Logging ##################")
elif args.verbose == 2:
    logging.basicConfig(level=logging.DEBUG)
    logging.debug("################## Starting - With verbose Logging ##################")
else: print("##################      Starting      ##################")

#################################################################################################
# ARG Parsing and Sorting                                                                       #
#################################################################################################

# arg api_server static or cli (cli overrides)
if not args.api_server is None:
    api_server = args.api_server
else:
    try: api_server
    except: api_server = input("Enter Check Point API Endpoint (Management Server / MaaS Instance):")
    if api_server == "": sys.exit("No API Server given!")

############################################################################
##### API Authentication #####
# check if api_user is given inside file or cli arguments. cli args override
if not args.api_user is None:
    api_user = args.api_user
else:
    try: api_user
    except: 
        if not args.key: 
            api_user = input("Enter Check Point API Username:")
        else: 
            api_user=""
            logging.warning("No API User given - expecting API Key authentication")

# check if api_pwd is given inside file or cli arguments. cli args override
if not args.api_pwd == None:
    api_pwd = args.api_pwd
else:
    try: api_pwd
    except: api_pwd=getpass.getpass(prompt="Enter Check Point API Password / API Key:")
if api_pwd == "": sys.exit("No password / API Key given - aborting!")

##### /API Authentication #####
############################################################################
##### Context (i.e. when using Management as a Service) ####
if not args.api_context == None:
    api_context = args.api_context
else: 
    try:
        api_context
    except: api_context = None

##### /Context (i.e. when using Management as a Service) ####
############################################################################
#### Command / What to Do ####
if not args.cmd == None:
    cmd = args.cmd
else:
    try:
        cmd
    except:
        sys.exit("No command given (i.e. show / export / test)")

if cmd in ["show","export"]:
    try: 
        choice = args.choice
    except: 
        sys.exit("No object type given (i.e. hosts, groups)")
    else: 
        choice = args.choice
        match choice:
            case "uo-content":
                apicommand = "show-updatable-objects-repository-content"
            case "uo-repos":
                apicommand = "show-updatable-objects"
            case "hosts":
                apicommand = "show-hosts"
            case "groups":
                apicommand = "show-groups"
            case "networks":
                apicommand = "show-networks"
            case "packages":
                apicommand = "show-packages"
            case "policy":
                apicommand = "show-access-rulebase"
                try: polname = args.polname
                except:
                    sys.exit("No policy given, but needed for this! Please add -p 'policy name' (i.e. 'Standard network')")
            case "unused-objects":
                # future use
                apicommand = "show-unused-objects"
            case "services":
                # future use
                apicommand = "show-services"
            case "zerohit":
                # future use
                # apicommand = str(cmd)+"-hosts"
                pass
            case other:
                sys.exit("unclear, what to do (variable choice undefined or incorrect)")
elif cmd == "test":
    apicommand = "show-session"
else: sys.exit("cmd not given or incorrect")
#### /Command aka. What to Do ####
############################################################################
#### Fields to use for export ####
# global exportselect
if not args.exportselect == None:
    exportselect = args.exportselect
else:
    try: exportselect
    except: exportselect = None
#### /Fields to use for export ####
############################################################################
#### Filter API response ####
if not args.cfilter == None:
    cfilter = args.cfilter
else: 
    try: cfilter
    except: cfilter = None
    else: pass

# Not sure if necessary - Warning that output might excessive...

# if not args.noninteractive and cfilter == None and not cmd == "export" and not cmd == "test": 
#     print("CAUTION : Without given filter, this script will try to get ALL objects! This might cause some load on server. Consider using a filter, like: -f \"tcp\" ")
#     cnt = input("Continue? y/n")
#     if cnt.lower()=="n" or cnt.lower()=="no":
#         sys.exit("Exiting as requested")
#     else: print("Continuing")

#### /Filter API response ####
############################################################################
#### Output file for export ####
if cmd == "export":
    if not args.outfile == None:
        outfile = args.outfile
    else:
        try: outfile
        except:
            outfile = input("Currently no destination file defined for exporting. Please set filename NOW - let empty for auto generated filename 'date_time_command.csv':")
            if outfile == "":
                try: 
                    polname
                except:
                    outfile = "csv-export_{}_{:%Y-%m-%d_%H-%M}.csv".format(apicommand,datetime.datetime.now())
                else:
                    outfile = "csv-export_{}_{:%Y-%m-%d_%H-%M}.csv".format(apicommand,datetime.datetime.now())
elif cmd == "show" and choice == "policy":
    sys.exit("Not recommended, as excessive output is possible - use 'export policy' instead")
elif cmd == "show": outfile=None
elif cmd == "test": outfile=None

#################################################################################################
# FUNCTIONS                                                                                     #
#################################################################################################
def fun_build_request():
    match apicommand:
        case "show-hosts":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"true"}
        case "show-networks":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"true"}
        case "show-groups":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"True",
                           "dereference-group-members":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"true",
                           "dereference-group-members":"True"}
        case "show-access-rulebase":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "name":polname,
                           "show-membership":"True",
                           "dereference-group-members":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "name":polname,
                           "show-membership":"true",
                           "dereference-group-members":"True"}
        case "show-services":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"true"}
        case "show-unused-objects":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"True",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full",
                           "show-membership":"true"}
        case "show-updatable-objects-repository-content":
            if filter:
                content = {"limit":100,
                           "details-level":"full",
                           "filter":cfilter}
            else:
                content = {"limit":100,
                           "details-level":"full"}
        case "test":
            content = {}
        case _:
            content = {}
    return(content)

def fun_apicomm(content,command):
    res_api = client.api_call(command, content)
    if not command == "show-session":
        if not res_api.success:
            logging.debug("FAILURE: ("+str(res_api)+")") 
            sys.exit("API Call not successful!"+str(res_api.data))
        else:
            return res_api.data
    else:
            return res_api
    
def fun_objectwork(objects, command, exportselect):
    if not exportselect == None:
        # checking wanted output values
        exportselect = exportselect.replace(" ","").split(",")
        for value in exportselect:
            match command:
                # check, if wanted fields exist in output
                case "show-hosts":
                    if not value in ("name",
                                    "type",
                                    "ipv4-address",
                                    "groups",
                                    "icon",
                                    "color",
                                    "comments",
                                    "uid",
                                    "tags",
                                    "last-modify-time",
                                    "last-modifier",
                                    "creation-time",
                                    "creator"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass
                case "show-networks":
                    if not value in ("name",
                                    "type",
                                    "subnet",
                                    "subnet-mask",
                                    "mask-length4",
                                    "groups",
                                    "icon",
                                    "color",
                                    "comments",
                                    "uid",
                                    "tags",
                                    "last-modify-time",
                                    "last-modifier",
                                    "creation-time","creator"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass
                case "show-groups": 
                    if not value in ("name",
                                    "type",
                                    "groups",
                                    "icon",
                                    "color",
                                    "comments",
                                    "uid",
                                    "tags",
                                    "last-modify-time",
                                    "last-modifier",
                                    "creation-time",
                                    "creator"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass
                case "show-unused-objects":
                    if not value in ("name",
                                    "type",
                                    "groups",
                                    "icon",
                                    "color",
                                    "comments",
                                    "uid",
                                    "tags",
                                    "last-modify-time",
                                    "last-modifier",
                                    "creation-time",
                                    "creator"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass
                case command if command.startswith("show-service"): 
                    if not value in ("name",
                                    "type",
                                    "groups",
                                    "icon",
                                    "color",
                                    "comments",
                                    "uid",
                                    "tags",
                                    "last-modify-time",
                                    "last-modifier",
                                    "creation-time",
                                    "creator"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass
                case command if command.startswith("show-updatable-objects"): 
                    if not value in ("name-in-updatable-objects-repository",
                                    "uri",
                                    "info-text",
                                    "description",
                                    "info-url",
                                    "uid-in-updatable-objects-repository"):
                        print("ERROR with exportselect argument {value}, please correct - may be ('name', 'type', 'ipv4-address','groups','icon','color','comments', 'uid', 'tags','last-modify-time', 'last-modifier', 'creation-time','creator')")
                    else:
                        pass

        fieldnames = exportselect
    else: 
        # default tables
        match command:
            case "show-hosts":
                fieldnames = ["name",
                              "type",
                              "ipv4-address",
                              "groups",
                              "icon",
                              "color",
                              "comments",
                              "tags",
                              "uid"]
            case "show-networks":
                fieldnames = ["name",
                            "type",
                            "subnet","subnet-mask","groups","icon","color","comments","tags",
                            "uid"]
            case "show-groups":
                fieldnames = ["name",
                            "type",
                            "groups",
                            "members",
                            "icon",
                            "color",
                            "comments","tags",
                            "uid"]
            case "show-unused-objects":
                fieldnames = ["name",
                            "type",
                            "groups",
                            "icon",
                            "color",
                            "comments",
                            "tags",
                            "uid",
                            "last-modify-time",
                            "last-modifier",
                            "creation-time",
                            "creator"]
            case command if command.startswith("show-service"):
                fieldnames = ["name",
                            "type",
                            "port",
                            "groups",
                            "match-for-any",
                            "session-timeout",
                            "icon",
                            "color",
                            "comments",
                            "tags",
                            "uid"]
            case command if command.startswith("show-updatable-objects"):
                fieldnames = ["name-in-updatable-objects-repository",
                            "uri",
                            "info-text",
                            "description",
                            "info-url",
                            "uid-in-updatable-objects-repository"]

    for object in objects:
        row=[]
        try: rows 
        except: 
            rows=[]
            rows.append(fieldnames)
        for field in fieldnames:
            match field:
                case "groups":
                    groups=[]
                    for group in object["groups"]:
                        groups.append(group["name"])
                    else:row.append(groups)
                case "tags":
                    tags=[]
                    for tag in object["tags"]:
                        tags.append(tag["name"])
                    else:row.append(tags)
# Kind of special attributes (nested or not uniquely o be preseen (ipv4 vs ipv6 objects)
                case "last-modify-time":
                    row.append(object["meta-info"]["last-modify-time"]["iso-8601"])
                case "last-modifier":
                    row.append(object["meta-info"]["last-modifier"])
                case "creation-time":
                    row.append(object["meta-info"]["creation-time"]["iso-8601"])    
                case "creator":
                    row.append(object["meta-info"]["creator"])
                case "subnet":
                    try:
                        object["subnet4"]
                    except:
                        try:
                            object["subnet6"]
                        except:
                            logging.debug("FAILURE: configured to use subnet fields, but none found")
                            raise SystemError("FAILURE: configured to use subnet fields, but none found")
                        else:
                            row.append(object["subnet6"])
                    else:
                        row.append(object["subnet4"])
                case "subnet-mask":
                    try:
                        object["mask-length4"]
                    except:
                        try:
                            object["mask-length6"]
                        except:
                            logging.debug("FAILURE: configured to use subnet mask / length fields, but none found")
                            raise SystemError("FAILURE: configured to use subnet mask / length fields, but none found")
                        else:
                            row.append(object["mask-length6"])
                    else:
                        row.append(object["mask-length4"])
                case "members":
                    members=[]
                    for member in object["members"]:
                        members.append(member["name"])
                    else:row.append(members)
                case _: 
                    try:
                        object[field]
                    except:
                        if command.startswith("show-updatable-objects"):
                            try:
                                object["additional-properties"][field]
                            except:
                                pass
                            else:
                                row.append(object["additional-properties"][field])
                        else:
                            row.append("N/A")
                    else:
                        row.append(object[field])
        rows.append(row)
    return rows



def fun_policywork(rulebase,policy,objects):
    f=[]
    for r in rulebase:
# r section or rule?
        if r["type"]=="access-section":
            f=["section","",r["name"],"","","","","","","","","","","","","","","","","",len(r["rulebase"])]
            policy.append(f)
            rbase=r["rulebase"]

        if r["type"]=="access-rule":
            rbase=r
########
        for rule in rbase:
            f=fun_readpolicy(rule, objects)
            policy.append(f)
            try:
                rule["inline-layer"]
            except:
                pass
            else:
                for a in objects:
                    inlinecontent = {"details-level":"full"}
                    if a["uid"]==rule["inline-layer"]:
                        inlinecontent["name"]=a["name"]
                        break
                inlinebase=fun_apicomm(inlinecontent,"show-access-rulebase")
                il_rulebase=inlinebase["rulebase"]
                for items in inlinebase["objects-dictionary"]:
                        objects.append(items)
                while not inlinebase["to"] == inlinebase["total"]:
                    inlinecontent["offset"]=inlinebase["to"]
                    inlinebase=fun_apicomm(inlinecontent,"show-access-rulebase")
                    for items in inlinebase["rulebase"]:
                        il_rulebase.append(items)
                    for items in inlinebase["objects-dictionary"]:
                        objects.append(items)
                            
                for r in il_rulebase:
                    if r["type"]=="access-section":
                        f=["section","",r["name"],"","","","","","","","","","","","","","","","","","Length:"+str(len(r["rulebase"]))]
                        policy.append(f)
                        for subrule in r["rulebase"]:
                            subrule["rule-number"]=str(rule["rule-number"])+"."+str(subrule["rule-number"])
                            f=fun_readpolicy(subrule, objects)
                            policy.append(f)
                    elif r["type"]=="access-rule":
                        subrule=r
                        subrule["rule-number"]=str(rule["rule-number"])+"."+str(subrule["rule-number"])
                        f=fun_readpolicy(subrule, objects)
                        policy.append(f)

def fun_readpolicy(rule, objects):
    try:
        rulename=rule["name"]
    except:
        rulename=""
    try:
        comments=rule["comments"]
    except:    
        comments=""
    source=[]
    destination=[]
    service=[]
    content=[]
    vpn=[]
    action=[]
    track=[]
    time=[]
    poltarget=[]
    enabled=""
    ruletags=[]

    for s in rule["source"]:
        for o in objects:
            if o["uid"] == s:
                source.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["destination"]:
        for o in objects:
            if o["uid"] == s:
                destination.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["service"]:
        for o in objects:
            if o["uid"] == s:
                service.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["content"]:
        for o in objects:
            if o["uid"] == s:
                content.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))  
    for s in rule["vpn"]:
        for o in objects:
            if o["uid"] == s:
                vpn.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))           
    if rule["action"]:
        s = rule["action"]
        for o in objects:
            if o["uid"] == s:
                action.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    if rule["track"]:
        s = rule["track"]["type"]
        for o in objects:
            if o["uid"] == s:
                track.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["time"]:
        for o in objects:
            if o["uid"] == s:
                time.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["install-on"]:
        for o in objects:
            if o["uid"] == s:
                poltarget.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["tags"]:
        for o in objects:
            if o["uid"] == s:
                poltarget.append(o["name"])
                break
            #else: raise SystemError("Object {0} not found in object dictionary".format(s))
    if rule["enabled"]:
        enabled = rule["enabled"]
    else:
        enabled = False
    if rule["source-negate"]: source = "NOT! - "+str(source)
    if rule["destination-negate"]: destination = "NOT! - "+str(destination)
    if rule["service-negate"]: service = "NOT! - "+str(service)
    if rule["content-negate"]: content = "NOT! - "+str(content)
    f=[rule["rule-number"],enabled,rulename,"\n".join([ii for ii in source]),"\n".join([ii for ii in destination]),"\n".join([ii for ii in service]),content,time,vpn,action,track,"\n".join([ii for ii in poltarget]),comments,rule["custom-fields"]["field-1"],rule["custom-fields"]["field-2"],rule["custom-fields"]["field-3"],ruletags,rule["meta-info"]["creator"],rule["meta-info"]["last-modify-time"]["iso-8601"],rule["meta-info"]["last-modifier"],rule["uid"]]
    return f

def fun_writepolicy(policy, objects):
    fieldnames=["nr.", "enabled","rule name", "source","destination","service","content","time","vpn","action","track/logging","policy target","comments","custom-1","custom-2","custom-3","tags","creator","lastmodified","lastmodifier","ruleid"]
    for i, object in enumerate(objects):

# objects
        try: 
            object["groups"]
        except KeyError:
            pass
        else: 
            groups = list()
            for c in object["groups"]:
                groups.append(c["name"])
            objects[i]["groups"] = "\n".join([ii for ii in groups]) #groups
            groups = list()

        try: 
            object["members"]
        except KeyError:
            pass
        else: 
            members = list()
            try:
                object["members"]["name"]
            except:
                try:
                    object["members"][0]["name"]
                except:
                    for c in object["members"]:
                        try:
                            for o in objects:
                                if o["uid"] == c: 
                                    members.append(o["name"])
                        except KeyError:
                            members.append("unknown")
                else:
                    for c in object["members"]:
                        members.append(c["name"])
                objects[i]["members"] = "\n".join([ii for ii in members]) 
            else:
                objects[i]["members"] = members[i]["members"]["name"]

# acrole
        try: 
            object["networks"]
        except KeyError:
            pass
        else: 
            networks = list()
            if len(object) == 1:
                objects[i]["networks"] = objects[i]["networks"]["name"]
            elif object["networks"]=="any":
                networks="any"
            else:
                for c in object["networks"]:
                   networks.append(c["name"])
            objects[i]["networks"] = "\n".join([ii for ii in networks]) #exclude
            networks = list()


        try: 
            object["users"]
        except KeyError:
            pass
        else: 
            users = list()
            if len(object) == 1:
                objects[i]["users"] = objects[i]["users"]["name"]
            elif object["users"]=="any":
                users="any"
            else:
                for c in object["users"]:
                    users.append(c["name"])
            objects[i]["users"] = "\n".join([ii for ii in users]) #exclude
            users = list()

        try: 
            object["machines"]
        except KeyError:
            pass
        else: 
            machines = list()
            if len(object) == 1:
                objects[i]["machines"] = objects[i]["machines"]["name"]
            elif object["machines"]=="any":
                machines="any"
            else:
                for c in object["machines"]:
                    machines.append(c["name"])
                objects[i]["machines"] = "\n".join([ii for ii in machines]) #exclude
                machines = list()


        try: 
            object["remote-access-client"]
        except KeyError:
            pass
        else: 
            clients = list()
            try:
                objects[i]["remote-access-client"]["name"]
            except KeyError:
                for c in object["remote-access-client"]:
                    clients.append(c["name"])
                objects[i]["remote-access-client"] = "\n".join([ii for ii in clients]) 
            else:
                objects[i]["remote-access-client"] = objects[i]["remote-access-client"]["name"]
            objects[i]["remote-access-client"] = "\n".join([ii for ii in clients]) 
            clients = list()

# groups with exclusions
        try: 
            object["include"]
        except KeyError:
            pass
        else: 
            include = list()
            try:
                objects[i]["include"]["name"]
            except KeyError:
                for c in object["include"]:
                    include.append(c["name"])
                objects[i]["include"] = "\n".join([ii for ii in include]) 
            else:
                include.append(object["include"]["name"])
            objects[i]["include"] = "\n".join([ii for ii in include]) 
            include = list()


        try: 
            object["except"]
        except KeyError:
            pass
        else: 
            exclude = list()
            try:
                objects[i]["except"]["name"]
            except KeyError:
                for c in object["except"]:
                    exclude.append(c["name"])
                objects[i]["except"] = "\n".join([ii for ii in exclude]) 
            else:
                exclude.append(objects[i]["except"]["name"])
            objects[i]["except"] = "\n".join([ii for ii in exclude]) 
            exclude = list()




# tags
        try: 
            object["tags"]
        except KeyError:
            pass
        else: 
            tags = list()
            if len(object) == 1:
                objects[i]["tags"] = objects[i]["tags"]["name"]
            for c in object["tags"]:
                tags.append(c["name"])
            objects[i]["tags"] = "\n".join([ii for ii in tags]) #tags
            tags = list()

# data-center-queries

        try: 
            object["data-centers"]
        except KeyError:
            pass
        else: 
            datacenters = list()
            try:
                objects[i]["data-centers"]["name"]
            except:
                for c in object["data-centers"]:
                    datacenters.append(c["name"])
                objects[i]["data-centers"] = "\n".join([ii for ii in datacenters]) 
            else:
                objects[i]["data-centers"] = objects[i]["data-centers"]["name"]
            objects[i]["data-centers"] = "\n".join([ii for ii in datacenters]) 
            datacenters = list()


    if "pandas" in sys.modules and outfile.__contains__("xls"):
        dfpolicy = pandas.DataFrame(policy,columns=fieldnames)
        dfobjects = pandas.DataFrame(objects).drop_duplicates(subset="uid", keep="first")
        l=list()
        for o in objects:
            # create list of types
            if not o["type"] in l: l.append(o["type"])

        with pandas.ExcelWriter(    outfile,
                                    engine="openpyxl") as writer:
            dfpolicy.to_excel(writer, index = False, sheet_name=args.polname)
            sheet = writer.book[args.polname]
            sheet.auto_filter.ref = "$A$1:$"+str(get_column_letter(sheet.max_column))+"$"+str(sheet.max_row)
            # Style auf Spalte A anwenden
            for cell in sheet["A"]:
                cell.alignment = openpyxl.styles.Alignment(horizontal="right")
            
            maxlen=0
            i=1
            while i < sheet.max_column:
                ii = 1
                while ii < sheet.max_row:
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                    if len(str(sheet.cell(row=ii, column=i).internal_value)) > maxlen:
                        maxlen=len(str(sheet.cell(row=ii, column=i).internal_value))
                        sheet.cell(row=ii, column=i).alignment = Alignment(wrap_text=True,vertical="top") 
                    ii = ii+1
                i=i+1
                sheet.column_dimensions[get_column_letter(i)].width =  maxlen


            # formatting section lines
            fsectionrow = PatternFill(fill_type="solid", bgColor="FDFD96") 
            rulesection = Rule(type="expression", dxf=DifferentialStyle(fill=fsectionrow), stopIfTrue=False)
            rulesection.formula= ["$A2='section'"] 
            sheet.conditional_formatting.add("$A$2:$"+str(get_column_letter(sheet.max_column))+"$"+str(sheet.max_row), rulesection)

            # formatting disabled rules
            fdisabledrow = PatternFill(fill_type="solid", bgColor="C3C3C3")
            ruledisabled = Rule(type="expression", dxf=DifferentialStyle(fill=fdisabledrow), stopIfTrue=False)
            ruledisabled.formula= ["$B2=FALSE"]
            sheet.conditional_formatting.add("$A$2:$"+str(get_column_letter(sheet.max_column))+"$"+str(sheet.max_row), ruledisabled)

            for type in l:
                match type:
                    case "network":
                        fieldnames = ["name",
                                      "subnet4",
                                      "subnet-mask",
                                      "comments",
                                      "groups",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(sheet.max_column))+"$"+str(sheet.max_row)

                    case "host":
                        fieldnames = ["name",
                                      "ipv4-address",
                                      "host-servers",
                                      "comments",
                                      "groups",
                                      "nat-settings",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)

                    case "address-range":
                        fieldnames = ["name",
                                      "ipv4-address-first",
                                      "ipv4-address-last", 
                                      "comments", 
                                      "color", 
                                      "uid", 
                                      "tags", 
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "group":
                        fieldnames = ["name", 
                                      "members", 
                                      "groups", 
                                      "comments", 
                                      "color", 
                                      "uid", 
                                      "tags", 
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "group-with-exclusion":
                        fieldnames = ["name", 
                                      "include", 
                                      "except", 
                                      "comments", 
                                      "color", 
                                      "uid", 
                                      "tags", 
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "data-centers":
                        fieldnames = ["name", 
                                      "members", 
                                      "groups", 
                                      "comments", 
                                      "color", 
                                      "uid", 
                                      "tags", 
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "data-center-query":
                        fieldnames = ["name", 
                                      "data-centers", 
                                      "query-rules", 
                                      "using-all-data-center", 
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "data-center-object":
                        fieldnames = ["name",
                                      "name-in-data-center",
                                      "uid-in-data-center",
                                      "type-in-data-center",
                                      "data-center",
                                      "additional-properties",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "updatable-object":
                        fieldnames = ["name",
                                      "additional-properties",
                                      "name-in-updatable-objects-repository",
                                      "uid-in-updatable-objects-repository",
                                      "updatable-object-meta-info",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "application-site":
                        fieldnames = ["name",
                                      "application-id",
                                      "primary-category",
                                      "risk",
                                      "primary-category-id",
                                      "description",
                                      "url-list",
                                      "urls-defined-as-regular-expression",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"] 
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "application-site-category":
                        fieldnames = ["name",
                                      "description",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "application-site-group":
                        fieldnames = ["name",
                                      "members",
                                      "description",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "access-layer":
                        fieldnames = ["name",
                                      "firewall",
                                      "mobile-access",
                                      "applications-and-url-filtering",
                                      "content-awareness",
                                      "implicit-cleanup-action",
                                      "detect-using-x-forward-for",
                                      "shared",
                                      "dynamic-layer",
                                      "parent-layer", 
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"] 
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "service-tcp":
                        fieldnames = ["name",
                                      "port", 
                                      "groups",
                                      "match-by-protocol-signature",
                                      "override-default-settings",
                                      "session-timeout", 
                                      "use-default-session-timeout",
                                      "match-for-any",
                                      "sync-connections-on-cluster",
                                      "aggressive-aging",
                                      "keep-connections-open-after-policy-installation",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "service-udp":
                        fieldnames = ["name",
                                      "port",
                                      "groups",
                                      "match-by-protocol-signature",
                                      "override-default-settings",
                                      "session-timeout",
                                      "use-default-session-timeout",
                                      "match-for-any",
                                      "sync-connections-on-cluster",
                                      "aggressive-aging",
                                      "keep-connections-open-after-policy-installation",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "service-other":
                        fieldnames = ["name",
                                      "port", 
                                      "groups", 
                                      "match-by-protocol-signature",
                                      "override-default-settings", 
                                      "session-timeout",
                                      "use-default-session-timeout",
                                      "match-for-any",
                                      "sync-connections-on-cluster",
                                      "aggressive-aging",
                                      "keep-connections-open-after-policy-installation",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "service-group":
                        fieldnames = ["name",
                                      "members",
                                      "groups", 
                                      "comments", 
                                      "color", 
                                      "uid", 
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "simple-gateway":
                        fieldnames = ["name",
                                      "ipv4-address",
                                      "interfaces",
                                      "groups",
                                      "platform",
                                      "version",
                                      "hardware",
                                      "https-inspection",
                                      "vpn-settings",
                                      "sic-name",
                                      "sic-state",
                                      "trust-state",
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    case "access-role":
                        fieldnames = ["name",
                                      "networks",
                                      "users", 
                                      "machines", 
                                      "remote-access-client", 
                                      "comments",
                                      "color",
                                      "uid",
                                      "tags",
                                      "meta-info"]
                        dfobjects[dfobjects["type"] == type].to_excel(writer, index = False, sheet_name=type, columns=fieldnames)
                        writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)
                        
                        
                    
                #dfobjects.to_excel(writer, index = False, sheet_name=type)
            dfobjects.to_excel(writer, index = False, sheet_name="all_objects")
            writer.book[type].auto_filter.ref = "$A$1:$"+str(get_column_letter(writer.book[type].max_column))+"$"+str(writer.book[type].max_row)

            # book = writer.book
            # sheet = book.active
            #format_sectionrow = sheet.add_format({"bg_color": "#FFC7CE","font_color": "#9C0006"})



    elif outfile.__contains__(".csv"):
        with open(outfile, "w", newline="") as output:
            writer = csv.writer(output, delimiter=";", quotechar="\"", quoting=csv.QUOTE_MINIMAL)
            fieldnames=["nr.", "enabled","rule name", "source","destination","service","content","time","vpn","action","track/logging","policy target","comments","custom-1","custom-2","custom-3","tags","creator","lastmodified","lastmodifier","ruleid"]
            writer.writerow(fieldnames)
            writer.writerow(policy)
        print("Information have been written to {0} - DONE!".format(outfile))
    else:
        print("")
    
def fun_printresult(result):
    match apicommand:
        case "show-packages":
            print(pandas.DataFrame(result))
        case "show-session":
            print("""
                Connection state:\t\tSuccess:{0} ({1})
                Server:\t\t\t\t{2}
                User:\t\t\t\t{3}
                From:\t\t\t\t{4}
                """.format(result.success,result.status_code,result.data["connected-server"]["name"],result.data["user-name"], result.data["ip-address"]))
        
def fun_writeobjects(rows):
    if not outfile == None:
        with open(outfile, "w", newline="") as output:
            writer = csv.writer(output, delimiter=";", quotechar="\"", quoting=csv.QUOTE_MINIMAL)
            for row in rows:
                writer.writerow(row)
        print("Information have been written to {0} - DONE!".format(outfile))
    else:
        if pds == True:
            print(pandas.DataFrame(rows[1:],columns=rows[0]))
        else:
            for row in rows:
                print(row)


###########################################################################################
# Functions
if __name__ == "__main__":
    if api_context == None:
        client_args = APIClientArgs(server=api_server, unsafe="True")
    else:
        client_args = APIClientArgs(server=api_server,context=api_context, unsafe="True")
    with APIClient(client_args) as client:
        if client.check_fingerprint() is False:
            logging.debug("Logging into SMS not successful! Please troubleshoot/debug script! ")
            sys.exit("Something is Wrong - checking fingerprint")
        if api_user and api_pwd:
            try: login_res = client.login(api_user, api_pwd)
            except: sys.exit("Login not possible! Please check arguments / server to connect / connectivity!")
            if not login_res.success:
                sys.exit("Login not possible")
            else:
                logging.debug("OK! Logged on")
        elif api_pwd:
            try: login_res = client.login_with_api_key(api_pwd)
            except: sys.exit("Login not possible! Please check arguments / server to connect / connectivity!")
            if not login_res.success:
                sys.exit("Login not possible")
            else:
                logging.debug("OK! Logged on")
        else:
            sys.exit("no Login Creds provided")
        content=fun_build_request()
        if apicommand == "show-access-rulebase":
            tmp=fun_apicomm(content,apicommand)
            rulebase=tmp["rulebase"]
            objects=tmp["objects-dictionary"]
            while not tmp["to"] == tmp["total"]:
                content["offset"]=tmp["to"]
                tmp=fun_apicomm(content,apicommand)
                for items in tmp["rulebase"]:
                    rulebase.append(items)
                for items in tmp["objects-dictionary"]:
                    objects.append(items)
            policy=[]
            fun_policywork(rulebase,policy,objects)
            fun_writepolicy(policy, objects)
        elif apicommand == "show-packages":
            fun_printresult(fun_apicomm(content,apicommand)["packages"])
        elif apicommand == "show-session":
            fun_printresult(fun_apicomm(content,apicommand))
        elif apicommand in supported_commands:
            if apicommand=="show-services" and not cfilter:
                apicommand=["show-services-tcp",
                            "show-services-udp",
                            "show-services-icmp",
                            "show-services-icmp6",
                            "show-services-sctp",
                            "show-services-other"]
            elif apicommand=="show-services":
                for filter in cfilter:
                    match filter:
                        case "tcp":
                            apicommand="show-services-tcp"
                        case "udp":
                            apicommand="show-services-udp"
                        case "icmp":
                            apicommand="show-services-icmp"
                        case "icmp6":
                            apicommand="show-services-icmp6"
                        case "sctp":
                            apicommand="show-services-sctp"
                        case "other":
                            apicommand="show-services-other"
            elif apicommand in supported_commands:
                content["offset"]=0
                tmp=fun_apicomm(content,apicommand)
                objects=tmp["objects"]
                if not len(objects) == 0:
                    while not tmp["to"] == tmp["total"]:
                        conttemp=content
                        conttemp["offset"]=tmp["to"]
                        tmp=fun_apicomm(conttemp,apicommand)
                        for o in tmp["objects"]:
                            objects.append(o)
                    fun_writeobjects(fun_objectwork(objects,apicommand,exportselect))
                else: print("empty Response for {0}".format(apicommand)) 
