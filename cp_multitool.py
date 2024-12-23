# !/usr/bin/env python3
"""Script to fetch objects, policies from CP Mgmt Server and putting them into a csv / xls file"""
# 2024 - DME
# Script to fetch objects, policies from CP Mgmt Server and putting them into a csv / xls file
# https://www.buymeacoffee.com/leinadeuntdomus
#################################################################################################
import sys
import argparse
import csv
import logging
import datetime
import getpass

try:
    import pandas
except ImportError:
    print("Module 'pandas' not found, therefore exports into .xls are not possible.")
else:
    import pandas
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.utils import get_column_letter
from cpapi import APIClient, APIClientArgs

#################################################################################################
#
# USAGE: execute script with python according to informations given by help. Possible usages:
# you can set environment variables in lines below.
# so you don´t have to type them everytime, you execute the script.
#
# (Sensitive data should not be saved (like password)
#
# python3.exe export-objects_subparser.py (--api_server 192.168.0.2
# can be declared by variable "api_server") --api_pwd <apikey> -f "IoT Next" show groups
# "show groups" is the commands here.
# This is executed as api call "show-groups" using "IoT Next" as filter
#
# see "supported commands", Readme.md or documentation to find out, what is currently implemented
#
#################################################################################################
# set args (will overriden by command)
# static environment variables
# api_server    = "cpmgmt"
# api_user      = "admin"
# api_pwd       = "superS3curePass"
# api_context   = None
# exportselect  = ""  #"name, uid, groups, members .... "
# cmd           = ""  #"test, show, export"
# choice        = "hosts"
# cfilter       = "Check Point"
#
#################################################################################################


supported_commands = [
    "test",
    "show-hosts",
    "show-groups",
    "show-networks",
    "show-host",
    "show-group",
    "show-network",
    "show-services",
    "show-access-rulebase",
    "show-packages",
    "show-package",
    "show-unused-objects",
    "show-updatable-objects-repository-content",
    "show-simple-cluster",
    "show-simple-gateway",
    "show-simple-clusters",
    "show-simple-gateways",
]

parser = argparse.ArgumentParser(epilog=""""
Small tool to show or export items or objects from Check Point Management Servers database
ike host, networks policy, packages and more.
More information in github repo:
https://github.com/leinadred/py_cpmultitool""")

parser.add_argument("-H", "--api_server", help="Target Host (CP Management Server)")
parser.add_argument("-U", "--api_user", help="API User")
parser.add_argument(
    "-K",
    "--key",
    help="invoke, that password value is an API Key",
    action="store_true",
    default=False,
)
parser.add_argument(
    "-P",
    "--api_pwd",
    help="API Users Password, if using API Key, use this without a user",
)

parser.add_argument(
    "-C",
    "--api_context",
    help="""If SmartCloud-1 is used, enter context information
    (i.e. bhkjnkm-knjhbas-d32424b/web_api) - defaults to '-C none'""",
)

parser.add_argument(
    "-p",
    "--polname",
    help='when using "show-access-rulebase" a policy name must be given, using this argument',
)

parser.add_argument(
    "-f", "--filter", dest="cfilter", help="filter by string (if applicable)"
)

parser.add_argument("-o", "--outfile", help="filename, to save the output in")
parser.add_argument(
    "-e",
    "--exportselect",
    help="""choose which fields to export to csv
    (define as string with quotes: "name, ip_address, uid, groups")""",
)
parser.add_argument(
    "-v",
    "--verbose",
    help="""Run Script with verbose logging output.
    For troubleshooting purposes.""",
    action="count",
)

subparser = parser.add_subparsers(
    required=True, dest="cmd", help="Tell what to do (test || show || export)"
)
parser_show = subparser.add_parser(
    "show",
    help="Print given information / objects and their properties (use with caution)",
)
parser_export = subparser.add_parser(
    "export",
    help="Save output of given information / objects and their properties to a file",
)
subparser.add_parser(
    "test",
    help="Basic connectivity test and (if successful) fetch some information from server",
)

parser_show.add_argument(
    dest="choice",
    choices=[
        "hosts",
        "host",
        "groups",
        "group",
        "networks",
        "network",
        "services",
        "policy",
        "unused-objects",
        "packages",
        "package",
        "uo-content",
        "clusters",
        "gateways",
        "cluster",
        "gateway",
    ],
)

parser_export.add_argument(
    dest="choice",
    choices=[
        "hosts",
        "host",
        "groups",
        "group",
        "networks",
        "network",
        "services",
        "policy",
        "unused-objects",
        "uo-content",
        "clusters",
        "gateways",
        "cluster",
        "gateway",
    ],
)

parser_show.add_argument(
    "objectname",
    type=str,
    nargs="?",
    default=None,
    help="""Name of Object to show (when used command is
    "show gateway", "show cluster", "show package")"""
)
parser_export.add_argument(
    "objectname",
    type=str,
    nargs="?",
    default=None,
    help="""Name of object to export (when using command is 'export policy',
    "export gateway", "show cluster", "show package") or name of access-rule-base.
    (when command is 'export policy')"""
)

args = parser.parse_args()


#################################################################################################
# ADDING DEBUG MODE                                                                             #
#################################################################################################
if args.verbose == 1:
    logging.basicConfig(level=logging.INFO)
    logging.debug(
        "################## Starting - With informational Logging ##################"
    )
elif args.verbose == 2:
    logging.basicConfig(level=logging.DEBUG)
    logging.debug(
        "################## Starting - With verbose Logging ##################"
    )
else:
    print("##################      Starting      ##################")

#################################################################################################
# ARG Parsing and Sorting                                                                       #
#################################################################################################

# arg api_server static or cli (cli overrides)
if not args.api_server is None:
    api_server = args.api_server
else:
    try:
        api_server
    except NameError:
        api_server = input(
            "Enter Check Point API Endpoint (Management Server / MaaS Instance):"
        )
    if api_server == "":
        sys.exit("No API Server given!")

############################################################################
##### API Authentication #####
# check if api_user is given inside file or cli arguments. cli args override
if not args.api_user is None:
    api_user = args.api_user
else:
    try:
        api_user
    except NameError:
        if not args.key:
            api_user = input("Enter Check Point API Username:")
        else:
            api_user = ""
            logging.info("No API User given - expecting API Key authentication")

# check if api_pwd is given inside file or cli arguments. cli args override
if not args.api_pwd is None:
    api_pwd = args.api_pwd
else:
    try:
        api_pwd
    except NameError:
        api_pwd = getpass.getpass(prompt="Enter Check Point API Password / API Key:")
if api_pwd == "":
    sys.exit("No password / API Key given - aborting!")

##### /API Authentication #####
############################################################################
##### Context (i.e. when using Management as a Service) ####
if not args.api_context is None:
    api_context = args.api_context
else:
    try:
        api_context  # type: ignore
    except NameError:
        api_context = None

##### /Context (i.e. when using Management as a Service) ####
############################################################################
#### Command / What to Do ####
if not args.cmd is None:
    cmd = args.cmd
else:
    try:
        cmd  # type: ignore
    except NameError:
        sys.exit("No command given (i.e. show / export / test)")

if cmd in ["show", "export"]:
    try:
        choice = args.choice
    except NameError:
        sys.exit("No object type given (i.e. hosts, groups)")
    else:
        choice = args.choice
        match choice:
            case "uo-content":
                apicommand = "show-updatable-objects-repository-content"
            case "uo-repos":
                apicommand = "show-updatable-objects"
            case "hosts":
                apicommand = "show-hosts"
            case "groups":
                apicommand = "show-groups"
            case "networks":
                apicommand = "show-networks"
            case "groups":
                apicommand = "show-groups"
            case "host":
                apicommand = "show-host"
            case "network":
                apicommand = "show-network"
            case "group":
                apicommand = "show-group"
            case "packages":
                apicommand = "show-packages"
            case "package":
                apicommand = "show-package"
            case "policy":
                apicommand = "show-access-rulebase"
                try:
                    polname = args.objectname
                except NameError:
                    sys.exit(
                        """No policy given, but needed for this!
                        Please add -p 'policy name' (i.e. 'Standard network')"""
                    )
            case "unused-objects":
                apicommand = "show-unused-objects"
            case "services":
                # future use
                apicommand = "show-services"
            case "zerohit":
                # future use
                # apicommand = str(cmd)+"-hosts"
                pass
            case "clusters":
                apicommand = "show-simple-clusters"
            case "gateways":
                apicommand = "show-simple-gateways"
            case "cluster":
                apicommand = "show-simple-cluster"
            case "gateway":
                apicommand = "show-simple-gateway"
            case other:
                sys.exit("unclear, what to do (variable choice undefined or incorrect)")
elif cmd == "test":
    apicommand = "show-session"
else:
    sys.exit("cmd not given or incorrect")

#### /Command aka. What to Do ####
############################################################################
#### Fields to use for export ####
# global exportselect
if not args.exportselect is None:
    exportselect = args.exportselect
else:
    try:
        exportselect  # type: ignore
    except NameError:
        exportselect = None
#### /Fields to use for export ####
############################################################################
#### Filter API response ####
if not args.cfilter is None:
    cfilter = args.cfilter
else:
    try:
        cfilter  # type: ignore
    except NameError:
        cfilter = None
    else:
        pass

# Not sure if necessary - Warning that output might excessive...

# if not args.noninteractive and cfilter == None and not cmd == "export" and not cmd == "test":
#     print(f"""CAUTION : Without given filter, this script will try to get ALL objects!
#     This might cause some load on server. Consider using a filter, like: -f \"tcp\" """)
#     cnt = input("Continue? y/n")
#     if cnt.lower()=="n" or cnt.lower()=="no":
#         sys.exit("Exiting as requested")
#     else: print("Continuing")

#### /Filter API response ####
############################################################################
#### Output file for export ####
if cmd == "export":
    if not args.outfile is None:
        outfile = args.outfile
    else:
        try:
            outfile  # type: ignore
        except NameError:
            outfile = input(
                """Currently no destination file defined for exporting. Please set filename NOW
                 let empty for auto generated filename 'date_time_command.csv':"""
            )
            if outfile == "":
                outfile = f"""
{datetime.datetime.now():%Y%m%d-%H%M}_{apicommand}-{polname.replace(" ","__")}"""
elif cmd == "show" and choice == "policy":
    sys.exit(
        "Not recommended, as excessive output is possible - use 'export policy' instead"
    )
elif cmd == "show":
    outfile = None
elif cmd == "test":
    outfile = None


#################################################################################################
# FUNCTIONS                                                                                     #
#################################################################################################
def fun_build_request():
    """Building API Request"""
    match apicommand:
        case "show-hosts":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "True",
                    "filter": cfilter,
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "true",
                }
        case "show-networks":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "True",
                    "filter": cfilter,
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "true",
                }
        case "show-groups":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "True",
                    "dereference-group-members": "True",
                    "filter": cfilter,
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "true",
                    "dereference-group-members": "True",
                }
        case "show-access-rulebase":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "name": polname,
                    "show-membership": "True",
                    "dereference-group-members": "True",
                    "filter": cfilter,
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "name": polname,
                    "show-membership": "true",
                    "dereference-group-members": "True",
                }
        case "show-services":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "True",
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "true",
                }
        case "show-unused-objects":
            if filter:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "True",
                    "filter": cfilter,
                }
            else:
                request = {
                    "limit": 100,
                    "details-level": "full",
                    "show-membership": "true",
                }
        case "show-updatable-objects-repository-content":
            if cfilter:
                request = {"limit": 100, "details-level": "full", "filter": cfilter}
            else:
                request = {"limit": 100, "details-level": "full"}
        case "show-package":
            request = {"name": args.objectname}
        case "test":
            request = {}
        case "show-simple-gateway":
            if args.objectname:
                request = {"details-level": "full", "name": args.objectname}
            elif cfilter:
                request = {"details-level": "full", "filter": cfilter}
        case "show-simple-cluster":
            if args.objectname:
                request = { "details-level": "full", "name": args.objectname}
            elif cfilter:
                request = { "details-level": "full", "filter": cfilter}

        case "show-host":
            if args.objectname:
                request = { "details-level": "full", "name": args.objectname}
            elif cfilter:
                request = { "details-level": "full", "filter": cfilter}

        case "show-network":
            if args.objectname:
                request = { "details-level": "full", "name": args.objectname}
            elif cfilter:
                request = { "details-level": "full", "filter": cfilter}

        case "show-group":
            if args.objectname:
                request = { "details-level": "full", "name": args.objectname}
            elif cfilter:
                request = { "details-level": "full", "filter": cfilter}

        case _:
            if cfilter:
                request = {"details-level": "full", "filter": cfilter}
            else:
                request = {"details-level": "full"}
    return request


def fun_apicomm(request, command):
    """Communication with API"""
    res_api = client.api_call(command, request)
    # if not command == "show-session":
    if not res_api.success:
        match res_api.status_code:
            case 400:
                logging.debug("FAILURE: (%s)", str(res_api))
                sys.exit(
                    f"""API Call not successful! Something wrong in request.
                    {str(res_api.data['message'])}
                    Please check again given arguments."""
                )
            case 401:
                logging.debug("FAILURE: (%s)", str(res_api))
                sys.exit(
                    f"""API Call not successful! Unauthorized
                    {str(res_api.data['message'])}
                    Please check login data (or change user)."""
                )
            case 404:
                logging.debug("FAILURE: (%s)", str(res_api))
                sys.exit(
                    f"""API Call not successful! Not Found
                    {str(res_api.data['message'])}
                    Please check given arguments."""
                )
            case 500:
                logging.debug("FAILURE: (%s)", str(res_api))
                sys.exit(
                    f"""API Call not successful! Internal Server Error
                    {str(res_api.data.message)}
                    Server error - check if API is enabled and troubleshoot API service."""
                )
            case _:
                logging.debug("FAILURE: (%s)", str(res_api))
                sys.exit(
                    f"""API Call not successful! Unknown error occured
                    {str(res_api.data)}
                    Please troubleshoot API service, check arguments."""
                )
    else:
        logging.debug("OK - Got response %s - Working!", res_api.data)
        logging.info("OK - Got status %s - Working!", res_api.status_code)
        return res_api


def fun_objectwork(objects, apicommand, exportselect):
    """working with extracted objects"""
    meta=[
            "last-modify-time",
            "last-modifier",
            "creation-time",
            "creator"
            ]

    if apicommand == "show-packages":
        objects = objects["packages"]
    try:
        objects[0]["uid"]
    except KeyError:
        try:
            objects["uid"]
        except KeyError as e:
            sys.exit("Failure %s in parsing output\n%s",e, objects)
        else:
            allowed = list(objects.keys()) + meta
    else:
        allowed = list(objects[0].keys()) + meta
    if not exportselect in [None,"all"]:
        # checking wanted output values
        fieldnames=[]
        exportselect = exportselect.replace(" ", "").split(",")
        for value in exportselect:
            if not value in allowed:
                sys.exit(
                    f"""ERROR with exportselect argument {value}, please correct!
                    May be ({allowed}) or "all" - example -e name,uid,creator
                    """
                )
            else:
                fieldnames.append(value)
    elif not exportselect == ['all']:
        fieldnames = allowed
    else:
        # default tables
        match apicommand:
            case "show-hosts":
                fieldnames = [
                    "name",
                    "type",
                    "ipv4-address",
                    "groups",
                    "icon",
                    "color",
                    "comments",
                    "tags",
                    "uid",
                ]
            case "show-networks":
                fieldnames = [
                    "name",
                    "type",
                    "subnet",
                    "subnet-mask",
                    "groups",
                    "icon",
                    "color",
                    "comments",
                    "tags",
                    "uid",
                ]
            case "show-groups":
                fieldnames = [
                    "name",
                    "type",
                    "groups",
                    "members",
                    "icon",
                    "color",
                    "comments",
                    "tags",
                    "uid",
                ]
            case "show-unused-objects":
                fieldnames = [
                    "name",
                    "type",
                    "groups",
                    "comments",
                    "tags",
                    "uid",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-service"):
                fieldnames = [
                    "name",
                    "type",
                    "port",
                    "groups",
                    "match-for-any",
                    "session-timeout",
                    "icon",
                    "color",
                    "comments",
                    "tags",
                    "uid",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-updatable-objects"):
                fieldnames = [
                    "name-in-updatable-objects-repository",
                    "uri",
                    "info-text",
                    "description",
                    "info-url",
                    "uid-in-updatable-objects-repository",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-simple-gateway"):
                fieldnames = [
                    "name",
                    "uid",
                    "type",
                    "platform",
                    "os-name",
                    "version",
                    "ipv4-address",
                    "sic-state",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-simple-cluster"):
                fieldnames = [
                    "name",
                    "uid",
                    "type",
                    "platform",
                    "os-name",
                    "version",
                    "ipv4-address",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-host"):
                fieldnames = [
                    "name",
                    "uid",
                    "type",
                    "comments",
                    "ipv4-address",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case apicommand if apicommand.startswith("show-group"):
                fieldnames = [
                    "name",
                    "uid",
                    "type",
                    "comments",
                    "members",
                    "last-modify-time",
                    "last-modifier",
                    "creation-time",
                    "creator",
                ]
            case _:
                fieldnames = allowed

    if apicommand in [
        "show-simple-gateway",
        "show-simple-cluster",
        "show-package",
        "show-host",
        "show-network",
        "show-group"
        ]:
        objects = [objects]
    for o in objects:
        row = []
        try:
            rows
        except NameError:
            rows = []
            rows.append(fieldnames)
        for field in fieldnames:
            match field:
                case "groups":
                    groups = []
                    for group in o["groups"]:
                        groups.append(group["name"])
                    row.append(groups)
                case "tags":
                    tags = []
                    for tag in o["tags"]:
                        tags.append(tag["name"])
                    row.append(tags)

                case "last-modify-time":
                    row.append(o["meta-info"]["last-modify-time"]["iso-8601"])
                case "last-modifier":
                    row.append(o["meta-info"]["last-modifier"])
                case "creation-time":
                    row.append(o["meta-info"]["creation-time"]["iso-8601"])
                case "creator":
                    row.append(o["meta-info"]["creator"])
                case "subnet":
                    try:
                        o["subnet4"]
                    except Exception:
                        try:
                            o["subnet6"]
                        except Exception as e:
                            logging.debug(
                                "FAILURE: configured to use subnet fields, but none found"
                            )
                            raise SystemError from e
                        else:
                            row.append(o["subnet6"])
                    else:
                        row.append(o["subnet4"])
                case "subnet-mask":
                    try:
                        o["mask-length4"]
                    except Exception as e:
                        try:
                            o["mask-length6"]
                        except:
                            logging.debug(
                                "FAILURE: no subnet mask / length fields found"
                            )
                            raise SystemError from e
                        else:
                            row.append(o["mask-length6"])
                    else:
                        row.append(o["mask-length4"])
                case "members":
                    members = []
                    for member in o["members"]:
                        members.append(member["name"])
                    row.append(members)
                case _:
                    try:
                        o[field]
                    except:
                        if apicommand.startswith("show-updatable-objects"):
                            try:
                                row.append(o["additional-properties"][field])
                            except:
                                logging.debug(
                                    "FAILURE: no additional-properties fields found"
                                )
                            else:
                                pass
                        else:
                            row.append("N/A")
                    else:
                        row.append(o[field])
        rows.append(row)
        
    logging.debug("OK - Work is done, moving on to create output")
    logging.info("OK - Work is done, moving on to create output (%s line/-s)",len(rows)-1)
    return rows, fieldnames


def fun_policywork(rulebase, policy, objects):
    """working with extracted policy"""
    f = []
    for line in rulebase:
        # r section or rule?
        if line["type"] == "access-section":
            f = [
                "section",
                "",
                line["name"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                len(line["rulebase"]),
            ]
            policy.append(f)
            try:
                rbase=line["rulebase"]
            except:
                logging.debug("%s seems to be an empty section", {line['name']})
                pass
            else:
                for rule in rbase:
                    f=fun_readpolicy(rule, objects)
                    policy.append(f)
                    try:
                        rule["inline-layer"]
                    except:
                        pass
                    else:
                        for a in objects:
                            inlinecontent = {"details-level":"full"}
                            if a["uid"]==rule["inline-layer"]:
                                inlinecontent["name"]=a["name"]
                                break
                        inlinebase=fun_apicomm(inlinecontent,"show-access-rulebase").data
                        il_rulebase=inlinebase["rulebase"]
                        for items in inlinebase["objects-dictionary"]:
                            objects.append(items)
                        while not inlinebase["to"] == inlinebase["total"]:
                            inlinecontent["offset"]=inlinebase["to"]
                            inlinebase=fun_apicomm(inlinecontent,"show-access-rulebase")
                            for items in inlinebase["rulebase"]:
                                il_rulebase.append(items)
                            for items in inlinebase["objects-dictionary"]:
                                objects.append(items)
                        for r in il_rulebase:
                            if r["type"]=="access-section":
                                f=[
                                    "section",
                                    "",
                                    r["name"],
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "Length:"+str(len(r["rulebase"]))
                                    ]
                                policy.append(f)
                                for subrule in r["rulebase"]:
                                    subrule["rule-number"]=f"""{rule["rule-number"]}.{subrule["rule-number"]}"""
                                    f=fun_readpolicy(
                                        subrule, 
                                        objects)
                                    policy.append(f)
                            elif r["type"]=="access-rule":
                                subrule=r
                                subrule["rule-number"]=f"""{rule["rule-number"]}.{subrule["rule-number"]}"""
                                f=fun_readpolicy(subrule, objects)
                                policy.append(f)

        elif line["type"] == "access-rule":
            f = fun_readpolicy(line, objects)
            policy.append(f)
            try:
                line["inline-layer"]
            except KeyError:
                pass
            else:
                for a in objects:
                    inlinecontent = {"details-level": "full"}
                    if a["uid"] == line["inline-layer"]:
                        inlinecontent["name"] = a["name"]
                        break
                inlinebase = fun_apicomm(inlinecontent, "show-access-rulebase").data
                il_rulebase = inlinebase["rulebase"]
                for items in inlinebase["objects-dictionary"]:
                    objects.append(items)
                while not inlinebase["to"] == inlinebase["total"]:
                    inlinecontent["offset"] = inlinebase["to"]
                    inlinebase = fun_apicomm(inlinecontent, "show-access-rulebase")
                    for items in inlinebase["rulebase"]:
                        il_rulebase.append(items)
                    for items in inlinebase["objects-dictionary"]:
                        objects.append(items)

                for r in il_rulebase:
                    if r["type"] == "access-section":
                        f = [
                            "section",
                            "",
                            r["name"],
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "Length:" + str(len(r["rulebase"])),
                        ]
                        policy.append(f)
                        for subrule in r["rulebase"]:
                            subrule["rule-number"] = (
                                str(rule["rule-number"])
                                + "."
                                + str(subrule["rule-number"])
                            )
                            f = fun_readpolicy(subrule, objects)
                            policy.append(f)
                    elif r["type"] == "access-rule":
                        subrule = r
                        subrule["rule-number"] = (
                            str(line["rule-number"]) + "." + str(subrule["rule-number"])
                        )
                        f = fun_readpolicy(subrule, objects)
                        policy.append(f)


def fun_readpolicy(rule, objects):
    """Read Policy and Objects"""
    try:
        rulename = rule["name"]
    except KeyError:
        rulename = ""
    try:
        comments = rule["comments"]
    except KeyError:
        comments = ""
    source = []
    destination = []
    service = []
    content = []
    vpn = []
    action = []
    track = []
    time = []
    poltarget = []
    enabled = ""
    ruletags = []

    for s in rule["source"]:
        for o in objects:
            if o["uid"] == s:
                source.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["destination"]:
        for o in objects:
            if o["uid"] == s:
                destination.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["service"]:
        for o in objects:
            if o["uid"] == s:
                service.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["content"]:
        for o in objects:
            if o["uid"] == s:
                content.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["vpn"]:
        for o in objects:
            if o["uid"] == s:
                vpn.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    if rule["action"]:
        s = rule["action"]
        for o in objects:
            if o["uid"] == s:
                action.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    if rule["track"]:
        s = rule["track"]["type"]
        for o in objects:
            if o["uid"] == s:
                track.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["time"]:
        for o in objects:
            if o["uid"] == s:
                time.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["install-on"]:
        for o in objects:
            if o["uid"] == s:
                poltarget.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    for s in rule["tags"]:
        for o in objects:
            if o["uid"] == s:
                poltarget.append(o["name"])
                break
            # else: raise SystemError("Object {0} not found in object dictionary".format(s))
    if rule["enabled"]:
        enabled = rule["enabled"]
    else:
        enabled = False
    if rule["source-negate"]:
        source = "NOT! - " + str(source)
    if rule["destination-negate"]:
        destination = "NOT! - " + str(destination)
    if rule["service-negate"]:
        service = "NOT! - " + str(service)
    if rule["content-negate"]:
        content = "NOT! - " + str(content)
    f = [
        rule["rule-number"],
        enabled,
        rulename,
        "\n".join([ii for ii in source]),
        "\n".join([ii for ii in destination]),
        "\n".join([ii for ii in service]),
        content,
        time,
        vpn,
        action,
        track,
        "\n".join([ii for ii in poltarget]),
        comments,
        rule["custom-fields"]["field-1"],
        rule["custom-fields"]["field-2"],
        rule["custom-fields"]["field-3"],
        ruletags,
        rule["meta-info"]["creator"],
        rule["meta-info"]["creation-time"]["iso-8601"],
        rule["meta-info"]["last-modifier"],
        rule["meta-info"]["last-modify-time"]["iso-8601"],
        rule["uid"],
    ]
    return f


def fun_writepolicy(policy, objects):
    """write policy to file /stdout"""
    fieldnames = [
        "nr.",
        "enabled",
        "rule name",
        "source",
        "destination",
        "service",
        "content",
        "time",
        "vpn",
        "action",
        "track/logging",
        "policy target",
        "comments",
        "custom-1",
        "custom-2",
        "custom-3",
        "tags",
        "creator",
        "creation-time",
        "lastmodifier",
        "lastmodified",
        "ruleid",
    ]
    for i, o in enumerate(objects):

# extracting groups attribute

        try:
            o["groups"]
        except KeyError:
            pass
        else:
            groups = list()
            for c in o["groups"]:
                groups.append(c["name"])
            objects[i]["groups"] = "\n".join([ii for ii in groups])  # groups
            groups = list()

# extracting members attribute

        try:
            o["members"]
        except KeyError:
            pass
        else:
            members = list()
            try:
                objects[i]["members"] = members[i]["members"]["name"]
            except:
                try:
                    o["members"][0]["name"]
                except TypeError:
                    for c in o["members"]:
                        try:
                            for o in objects:
                                if o["uid"] == c:
                                    members.append(o["name"])
                        except KeyError:
                            members.append("unknown")
                else:
                    for c in o["members"]:
                        members.append(c["name"])
                objects[i]["members"] = "\n".join([ii for ii in members])

# extracting networks from Access Role object

        try:
            o["networks"]
        except KeyError:
            pass
        else:
            networks = list()
            if len(o) == 1:
                objects[i]["networks"] = objects[i]["networks"]["name"]
            elif o["networks"] == "any":
                networks = "any"
            else:
                for c in o["networks"]:
                    networks.append(c["name"])
            objects[i]["networks"] = "\n".join([ii for ii in networks])  # exclude
            networks = list()

# extracting users from Access Role object

        try:
            o["users"]
        except KeyError:
            pass
        else:
            users = list()
            if len(o) == 1:
                objects[i]["users"] = objects[i]["users"]["name"]
            elif o["users"] == "any":
                users = "any"
            else:
                for c in o["users"]:
                    users.append(c["name"])
            objects[i]["users"] = "\n".join([ii for ii in users])  # exclude
            users = list()

# extracting machines from Access Role object

        try:
            o["machines"]
        except KeyError:
            pass
        else:
            machines = list()
            if len(o) == 1:
                objects[i]["machines"] = objects[i]["machines"]["name"]
            elif o["machines"] == "any":
                machines = "any"
            else:
                for c in o["machines"]:
                    machines.append(c["name"])
                objects[i]["machines"] = "\n".join([ii for ii in machines])  # exclude
                machines = list()

# extracting remote-access-client from Access Role object

        try:
            o["remote-access-client"]
        except KeyError:
            pass
        else:
            clients = list()
            try:
                objects[i]["remote-access-client"]["name"]
            except KeyError:
                for c in o["remote-access-client"]:
                    clients.append(c["name"])
                objects[i]["remote-access-client"] = "\n".join([ii for ii in clients])
            else:
                objects[i]["remote-access-client"] = objects[i]["remote-access-client"][
                    "name"
                ]
            objects[i]["remote-access-client"] = "\n".join([ii for ii in clients])
            clients = list()

# extracting groups with exclusions

        try:
            o["include"]
        except KeyError:
            pass
        else:
            include = list()
            try:
                objects[i]["include"]["name"]
            except KeyError:
                for c in o["include"]:
                    include.append(c["name"])
                objects[i]["include"] = "\n".join([ii for ii in include])
            else:
                include.append(o["include"]["name"])
            objects[i]["include"] = "\n".join([ii for ii in include])
            include = list()

        try:
            o["except"]
        except KeyError:
            pass
        else:
            exclude = list()
            try:
                objects[i]["except"]["name"]
            except KeyError:
                for c in o["except"]:
                    exclude.append(c["name"])
                objects[i]["except"] = "\n".join([ii for ii in exclude])
            else:
                exclude.append(objects[i]["except"]["name"])
            objects[i]["except"] = "\n".join([ii for ii in exclude])
            exclude = list()

# extracting tags

        try:
            o["tags"]
        except KeyError:
            pass
        else:
            tags = list()
            if len(o) == 1:
                objects[i]["tags"] = objects[i]["tags"]["name"]
            for c in o["tags"]:
                tags.append(c["name"])
            objects[i]["tags"] = "\n".join([ii for ii in tags])  # tags
            tags = list()

# extracting data-center-queries

        try:
            o["data-centers"]
        except KeyError:
            pass
        else:
            datacenters = list()
            try:
                objects[i]["data-centers"] = objects[i]["data-centers"]["name"]
            except TypeError:
                for c in o["data-centers"]:
                    datacenters.append(c["name"])
                objects[i]["data-centers"] = "\n".join([ii for ii in datacenters])
            else:
                objects[i]["data-centers"] = objects[i]["data-centers"]["name"]
            objects[i]["data-centers"] = "\n".join([ii for ii in datacenters])
            datacenters = list()

# extracting meta infos

        try:
            o["meta-info"]
        except KeyError:
            pass
        else:
            try:
                o["meta-info"]["validation-state"]
            except KeyError:
                pass
            else:
                try:
                    objects[i]["validation-state"] = o["meta-info"]["validation-state"]
                except Exception as e:
                    logging.debug("Exception while extracting validation-state:\n %s", e)

# extracting last-modified-time
            try:
                o["meta-info"]["last-modify-time"]
            except KeyError:
                pass
            else:
                try:
                    objects[i]["last-modify-time"] = o["meta-info"]["last-modify-time"]["iso-8601"]
                except Exception as e:
                    logging.debug("Exception while extracting last modified time:\n %s", e)

# extracting last-modifier
            try:
                o["meta-info"]["last-modifier"]
            except KeyError:
                pass
            else:
                try:
                    objects[i]["last-modifier"] = o["meta-info"]["last-modifier"]
                except Exception as e:
                    logging.debug("Exception while extracting last modifier:\n %s", e)

# extracting creation-time
            try:
                o["meta-info"]["creation-time"]
            except KeyError:
                pass
            else:
                try:
                    objects[i]["creation-time"] = o["meta-info"]["creation-time"]["iso-8601"]
                except Exception as e:
                    logging.debug("Exception while extracting creation time:\n %s", e)

# extracting creator
            try:
                o["meta-info"]["creator"]
            except KeyError:
                pass
            else:
                try:
                    objects[i]["creator"] = o["meta-info"]["creator"]
                except Exception as e:
                    logging.debug("Exception while extracting creator:\n %s", e)

    if "pandas" in sys.modules and "xls" in outfile:
        dfpolicy = pandas.DataFrame(policy, columns=fieldnames)
        dfobjects = pandas.DataFrame(objects).drop_duplicates(
            subset="uid", keep="first"
        )

        with pandas.ExcelWriter(outfile, engine="openpyxl") as writer:
            dfpolicy.to_excel(writer, index=False, sheet_name=polname)
            sheet = writer.book[polname]

            # Berechne den Bereich für den AutoFilter
            max_row = dfpolicy.shape[0]
            max_col = dfpolicy.shape[1]
            sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                #"$A$1:$"+str(get_column_letter(sheet.max_column))+"$" + str(sheet.max_row))

            # Style auf Spalte A anwenden
            for cell in sheet["A"]:
                cell.alignment = openpyxl.styles.Alignment(horizontal="right")

            for i in range(1, sheet.max_column + 1):
                maxlen = 0
                # Iteriere über alle Zeilen in der aktuellen Spalte
                for ii in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=ii, column=i)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    maxlen = max(maxlen, get_max_char_per_line(cell))
                # Setze die Spaltenbreite basierend auf der maximalen Zeichenanzahl pro Zeile
                sheet.column_dimensions[get_column_letter(i)].width = maxlen+5

            # formatting section lines
            fsectionrow = PatternFill(fill_type="solid", bgColor="ffff6347")
            rulesection = Rule(
                type="expression",
                dxf=DifferentialStyle(fill=fsectionrow),
                stopIfTrue=False
            )
            rulesection.formula = ["$A2='section'"]
            sheet.conditional_formatting.add(
                f"$A$2:${get_column_letter(max_col)}${max_row}",
                rulesection
                )
            # formatting disabled rules
            fdisabledrow = PatternFill(fill_type="solid", bgColor="ff949494")
            ruledisabled = Rule(
                type="expression",
                dxf=DifferentialStyle(fill=fdisabledrow),
                stopIfTrue=False,
            )
            ruledisabled.formula = ["$B2=FALSE"]
            sheet.conditional_formatting.add(
                f"$A$2:${get_column_letter(max_col)}${max_row}",ruledisabled)

            for a in dfobjects["type"].drop_duplicates().to_list():
                writer.book.create_sheet(a)
                sheet=writer.book[a]
                try:
                    dfobjects[dfobjects["type"] == a].dropna(axis=1, how='all').to_excel(
                                writer, index=False, sheet_name=a
                            )
                    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

                    # Style auf Spalte A anwenden
                    for cell in sheet["A"]:
                        cell.alignment = openpyxl.styles.Alignment(horizontal="right")

                    for i in range(1, sheet.max_column + 1):
                        maxlen = 0
                        # Iteriere über alle Zeilen in der aktuellen Spalte
                        for ii in range(1, sheet.max_row + 1):
                            cell = sheet.cell(row=ii, column=i)
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            maxlen = max(maxlen, get_max_char_per_line(cell))
                        # Setze die Spaltenbreite basierend auf der maximalen Zeichenanzahl
                        sheet.column_dimensions[get_column_letter(i)].width = maxlen
                except:
                    sys.exit("Problem during writing Worksheet for %s",a)

            dfobjects.dropna(axis=1, how='all').to_excel(
                writer,
                index=False,
                sheet_name="all_objects"
                )

            writer.book.save(outfile)

    elif "csv" in outfile:
        with open(outfile, "w", newline="", encoding="utf-8") as output:
            writer = csv.writer(
                output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            fieldnames = [
                "nr.",
                "enabled",
                "rule name",
                "source",
                "destination",
                "service",
                "content",
                "time",
                "vpn",
                "action",
                "track/logging",
                "policy target",
                "comments",
                "custom-1",
                "custom-2",
                "custom-3",
                "tags",
                "creator",
                "lastmodified",
                "lastmodifier",
                "ruleid",
            ]
            writer.writerow(fieldnames)
            writer.writerow(policy)
            print(f"Information have been written to {outfile} - DONE!")
        print(f"Information have been written to {outfile} - DONE!")
    else:
        print("")

def get_max_char_per_line(cell):
    """function used to calculate max width for a cell/columns"""
    if cell.value is None:
        return 0
    lines = str(cell.value).split('\n')
    return max(len(line) for line in lines)


def fun_printresult(result):
    """Function to print the result to stdout"""
    match apicommand:
        case "show-packages":
            if "pandas" in sys.modules:
                print(pandas.DataFrame(result))
            else:
                print("Following policy packages have been found:\n")
                for a in result:
                    print(f"""name: {a['name']}, uid: {a['uid']}, type: {a['type']}""")
        case "show-package":
            result = result.data
            targets = []
            for a in result["installation-targets"]:
                targets.append(a["name"])
            print(
                f"""Name:\t{result['name']}
Type:\t{result['type']}
Domain:\t{result['domain']['name']}
Policy Targets:\t{targets}
------------------------------------------------------------"""
            )
            try:
                result["threat-layers"]
            except NameError:
                pass
            else:
                print("Access Layers:")
                for al in result["access-layers"]:
                    print(
                        f"""
    Name:\t{al['name']} (UID:{al['uid']})
    Domain:\t{al['domain']['name']} (UID:{al['domain']['uid']})"""
                    )
            print("-----------")
            try:
                result["threat-layers"]
            except NameError:
                pass
            else:
                print("Threat Layers:")
                for al in result["threat-layers"]:
                    print(
                        f"""
    Name:\t{al['name']} (UID:{al['uid']})
    Domain:\t{al['domain']['name']} (UID:{al['domain']['uid']})"""
                    )
            print("-----------")
            try:
                result["https-inspection-layers"]
            except NameError:
                pass
            else:
                print(
                    f"""HTTPS Layers:

    Inbound:\t{result["https-inspection-layers"]["inbound-https-layer"]['name']} (UID: {result["https-inspection-layers"]["inbound-https-layer"]['uid']})
    Outbound:\t{result["https-inspection-layers"]["outbound-https-layer"]['name']} (UID: ){result["https-inspection-layers"]["outbound-https-layer"]['uid']}
"""
                )
                print("-----------")
        case "show-session":
            print(
                f"""
                Connection state:\t\tSuccess:{result.success} ({result.status_code})
                Server:\t\t\t\t{result.data["connected-server"]["name"]}
                User:\t\t\t\t{result.data["user-name"]}
                From:\t\t\t\t{result.data["ip-address"]}
                """
            )


def fun_writeobjects(rows, fieldnames):
    """Function to write the objects result to stdout"""
    if not outfile is None:
        if "xls" in outfile and "pandas" in sys.modules:
            with pandas.ExcelWriter(outfile, engine="openpyxl") as writer:
                pandas.DataFrame(rows[1:], columns=fieldnames).to_excel(
                    writer, index=False, sheet_name=args.choice
                )
                print(f"Information have been written to {outfile} - DONE!")
        else:
            with open(outfile, "w", newline="", encoding="utf-8") as output:
                writer = csv.writer(
                    output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )
                for row in rows:
                    writer.writerow(row)
        print(f"Information have been written to {outfile} - DONE!")
    else:
        if "pandas" in sys.modules:
            print(pandas.DataFrame(rows[1:], columns=fieldnames))
        else:
            for row in rows:
                print(row)


###########################################################################################

if __name__ == "__main__":
    if api_context is None:
        client_args = APIClientArgs(server=api_server, unsafe="True")
    else:
        client_args = APIClientArgs(
            server=api_server, context=api_context, unsafe="True"
        )
    with APIClient(client_args) as client:
        if client.check_fingerprint() is False:
            logging.debug(
                "Logging into SMS not successful! Please troubleshoot/debug script! "
            )
            sys.exit("Something is Wrong - checking fingerprint")
        if api_user and api_pwd:
            try:
                login_res = client.login(api_user, api_pwd)
            except OSError as e:
                sys.exit(
                    f"""Connection failed!
Error: {e.strerror}
Please check arguments / server to connect / connectivity!"""
                )
            if not login_res.success:
                sys.exit("Login not possible")
            else:
                logging.debug("OK! Logged on")
        elif api_pwd:
            try:
                login_res = client.login_with_api_key(api_pwd)
            except OSError as e:
                sys.exit(
                    f"""Connection failed!
Error: {e.strerror}
Please check arguments / server to connect / connectivity!"""
                )
            if not login_res.success:
                sys.exit("Login not possible")
            else:
                logging.debug("OK! Logged on")
        else:
            sys.exit("no Login informations provided")
        content = fun_build_request()
        if apicommand == "show-access-rulebase":
            tmp = fun_apicomm(content, apicommand).data
            rulebase = tmp["rulebase"]
            objects = tmp["objects-dictionary"]
            while not tmp["to"] == tmp["total"]:
                content["offset"] = tmp["to"]
                tmp = fun_apicomm(content, apicommand).data
                for items in tmp["rulebase"]:
                    rulebase.append(items)
                for items in tmp["objects-dictionary"]:
                    objects.append(items)
            policy = []
            fun_policywork(rulebase, policy, objects)
            fun_writepolicy(policy, objects)
        # elif apicommand == "show-packages":
        #     fun_printresult(fun_apicomm(content, apicommand).data["packages"])
        # elif apicommand == "show-package":
        #     fun_printresult(fun_apicomm(content, apicommand))
        elif apicommand == "show-session":
            fun_printresult(fun_apicomm(content, apicommand))
        elif apicommand in supported_commands:
            servicetypes = [
                "tcp",
                "udp",
                "icmp",
                "icmp6",
                "sctp",
                "other",
                "dce-rpc",
                "rpc",
                "gtp",
                "citrix-tcp",
                "compound-tcp",
            ]
            if apicommand == "show-services" and not cfilter:
                sys.exit(
                    f"""There was an issue with given filter, allowed options are: {servicetypes}.
                    Using a service type is mandatory"""
                )
            elif apicommand == "show-services":
                if isinstance(cfilter) == str and cfilter in servicetypes:
                    apicommand = str(apicommand + "-" + cfilter)
                else:
                    for filter in cfilter:
                        match filter:
                            case "tcp":
                                apicommand = "show-services-tcp"
                            case "udp":
                                apicommand = "show-services-udp"
                            case "icmp":
                                apicommand = "show-services-icmp"
                            case "icmp6":
                                apicommand = "show-services-icmp6"
                            case "sctp":
                                apicommand = "show-services-sctp"
                            case "other":
                                apicommand = "show-services-other"
                            case "dce-rpc":
                                apicommand = "show-services-dce-rpc"
                            case "rpc":
                                apicommand = "show-services-rpc"
                            case "gtp":
                                apicommand = "show-services-gtp"
                            case _:
                                sys.exit(
                                    f"""There was an issue with given filter!
                                         Allowed options are: {servicetypes}.
                                         Using a service type is mandatory"""
                                )
            elif apicommand in ["show-simple-gateway","show-simple-cluster","show-package"]:
                tmp = fun_apicomm(content, apicommand).data
                objects = tmp
                #objects = tmp["objects"]
                if not len(objects) == 0:
                    # while not tmp["to"] == tmp["total"]:
                    #     conttemp = content
                    #     conttemp["offset"] = tmp["to"]
                    #     tmp = fun_apicomm(conttemp, apicommand).data
                    #     for o in tmp["objects"]:
                    #         objects.append(o)
                    rows, fieldnames = fun_objectwork(objects, apicommand, exportselect)
                    fun_writeobjects(rows, fieldnames)
                else:
                    print(f"empty Response for {apicommand}")
            elif apicommand in ["show-host","show-network","show-group"]:
                tmp = fun_apicomm(content, apicommand).data
                objects = tmp
                rows, fieldnames = fun_objectwork(objects, apicommand, exportselect)
                fun_writeobjects(rows, fieldnames)
            elif apicommand in ["show-packages"]:
                tmp = fun_apicomm(content, apicommand).data
                objects = tmp
                rows, fieldnames = fun_objectwork(objects, apicommand, exportselect)
                fun_writeobjects(rows, fieldnames)
            else:
                tmp = fun_apicomm(content, apicommand).data
                objects = tmp["objects"]
                if not len(objects) == 0:
                    while not tmp["to"] == tmp["total"]:
                        conttemp = content
                        conttemp["offset"] = tmp["to"]
                        tmp = fun_apicomm(conttemp, apicommand).data
                        for o in tmp["objects"]:
                            objects.append(o)
                    rows, fieldnames = fun_objectwork(objects, apicommand, exportselect)
                    fun_writeobjects(rows, fieldnames)
                else:
                    sys.exit(f"empty Response for {apicommand}")
                # rows, fieldnames = fun_objectwork(objects, apicommand, exportselect)
                # fun_writeobjects(rows, fieldnames)
        else:
            sys.exit("Given arguments are not supported (yet). See Manual/Readme/help")
